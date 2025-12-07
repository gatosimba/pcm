# clinicas/admin.py
from pyexpat.errors      import messages
from django.contrib      import admin
from import_export.admin import ImportExportModelAdmin
from django              import forms
import                          openpyxl
from django.http         import HttpResponse

# Models da APP Clinicas
from clinicas.models     import *


#========================================
class ClinicaAdminBase(admin.ModelAdmin):
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs

        perfil = PerfilUsuario.objects.filter(user=request.user).first()
        if perfil and perfil.clinica:
            return qs.filter(clinica=perfil.clinica)
        return qs.none()

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "clinica":
            if request.user.is_superuser:
                return super().formfield_for_foreignkey(db_field, request, **kwargs)

            perfil = PerfilUsuario.objects.filter(user=request.user).first()
            if perfil and perfil.clinica:
                kwargs["queryset"] = Clinica.objects.filter(id=perfil.clinica.id)
                if not request.resolver_match.kwargs.get('object_id'):  # só no ADD
                    kwargs["initial"] = perfil.clinica
            else:
                kwargs["queryset"] = Clinica.objects.none()

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_readonly_fields(self, request, obj=None):
        # <<< LINHA CORRIGIDA >>>
        readonly = list(super().get_readonly_fields(request, obj) or ())
        if obj and not request.user.is_superuser:
            readonly.append('clinica')
        return readonly    







#====================================    
class ClinicaAdmin(admin.ModelAdmin):

    list_display = ('nome', 'responsavel', 'localidade', 'cep', 'status',) # 'user', 'data_criacao')
    #list_filter = ('status', 'uf')
    search_fields = ('nome', 'logradouro', 'bairro', 'localidade', 'cep')
    ordering = ('localidade', 'uf', 'nome',)

    # Impede que o usuário troque o "user" no admin
    readonly_fields = ('user', 'data_criacao', 'data_atualizacao')

    fieldsets = (
        ('Dados da Clínica', {
            'fields': (
                'nome',
                'responsavel',
                'status',
            )
        }),
        ('Endereço', {
            'fields': (
                'cep',
                'logradouro',
                'complemento',
                'bairro',
                'localidade',
                'uf',
            )
        }),
        ('Controle', {
            'fields': (
                'user',
                'data_criacao',
                'data_atualizacao',
            )
        }),
    )

    # --------------------------
    # Filtra lista por usuário
    # --------------------------
    def get_queryset(self, request):
        qs = super().get_queryset(request)

        if request.user.is_superuser:
            return qs  # vê tudo

        return qs.filter(user=request.user)   # só as clínicas dele

    # --------------------------
    # Ao criar uma clínica, vincula ao usuário automaticamente
    # --------------------------
    def save_model(self, request, obj, form, change):
        if not obj.pk:  # criando
            obj.user = request.user
        else:
            # impedimos troca manual de dono
            if not request.user.is_superuser:
                form.cleaned_data.pop('user', None)

        super().save_model(request, obj, form, change)

    # --------------------------
    # Função auxiliar p/ exibir cidade/estado
    # --------------------------
    def cidade_estado(self, obj):
        return f"{obj.localidade or ''}/{obj.uf or ''}"
    cidade_estado.short_description = 'Cidade/UF'

admin.site.register(Clinica, ClinicaAdmin)
#=========================================


# Expotar para Excel: Model Parametro
#====================================

def exportar_parametro_xlsx(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parametros"

    # Cabeçalho
    ws.append(["Codigo", "Valor", "Descrição", "Clínica"])

    # Linhas
    for obj in queryset:
        ws.append([obj.codigo, obj.valor, obj.descricao, obj.clinica.nome])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="parametros.xlsx"'

    wb.save(response)

    return response

exportar_parametro_xlsx.short_description = "Export Excel"
#================================================================


class ParametroAdmin(admin.ModelAdmin):

    list_display = ('codigo', 'valor', 'descricao', 'clinica',)
    search_fields = ('codigo', 'descricao', 'valor')
    #list_filter = ('clinica',)
    ordering = ('clinica__nome', 'codigo')

    actions = [exportar_parametro_xlsx]  # adiciona action normal
    change_list_template = "admin/parametros_xls.html"

# ---------------------------------------------
# 4. Exporta TODOS os registros da lista (sem precisar selecionar)
# ---------------------------------------------
    def get_actions(self, request):
        actions = super().get_actions(request)
        # Remove a action padrão do admin (que exige seleção)
        if 'exportar_parametro_xlsx' in actions:
            del actions['exportar_parametro_xlsx']
        return actions

    # Força a action a sempre usar todos os itens da queryset (a mágica!)
    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'exportar_parametro_xlsx':
            # Pega TODOS os objetos que o usuário pode ver (com filtro de clínica!)
            queryset = self.get_queryset(request)
            return exportar_parametro_xlsx(self, request, queryset)
        return super().changelist_view(request, extra_context)

    # Campos readonly para não permitir troca de dono
    readonly_fields = []

    fieldsets = (
        ('Identificação', {
            'fields': ('clinica', 'codigo')
        }),
        ('Informações', {
            'fields': ('valor', 'descricao')
        }),
    )

    # ---------------------------------------------
    # 1. Filtra lista para mostrar apenas da clínica do consultor
    # ---------------------------------------------
    def get_queryset(self, request):
        qs = super().get_queryset(request)

        if request.user.is_superuser:
            return qs

        return qs.filter(clinica__user=request.user)

    # ---------------------------------------------
    # 2. Filtra dropdown da clínica
    # ---------------------------------------------
    def formfield_for_foreignkey(self, db_field, request, **kwargs):

        if db_field.name == "clinica":
            if request.user.is_superuser:
                kwargs["queryset"] = Clinica.objects.all()
            else:
                kwargs["queryset"] = Clinica.objects.filter(user=request.user)

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # ---------------------------------------------
    # 3. Impede que o usuário troque a clínica depois que o registro existe
    # ---------------------------------------------
    def get_readonly_fields(self, request, obj=None):
        if obj:  # edição
            return ('clinica',)  # trava o campo
        return self.readonly_fields
#==================================
admin.site.register(Parametro, ParametroAdmin)
#=============================================


# Expotar para Excel: Model Categoria
#====================================

def exportar_categoria_xlsx(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorias"

    # Cabeçalho
    ws.append(["Codigo", "Descrição", "Tipo", "clinica"])

    # Linhas
    for obj in queryset:
        ws.append([obj.codigo, obj.descricao, obj.tipo, obj.clinica.nome])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="categorias.xlsx"'

    wb.save(response)

    return response

exportar_categoria_xlsx.short_description = "Export Excel"
#=========================================================

#@admin.register(Categoria)
class CategoriaAdmin(admin.ModelAdmin):
    list_display = ("codigo", "descricao", "tipo", "clinica")

    actions = [exportar_categoria_xlsx]  # adiciona action normal
    change_list_template = "admin/categorias_xls.html"

# ---------------------------------------------
# 4. Exporta TODOS os registros da lista (sem precisar selecionar)
# ---------------------------------------------
# 4. REMOVE A ACTION DO DROPDOWN (pra não confundir)
    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'exportar_categoria_xlsx' in actions:
            del actions['exportar_categoria_xlsx']
        return actions

# 5. A MÁGICA: clica no botão → exporta TODAS as linhas da lista
    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'exportar_categoria_xlsx':
            queryset = self.get_queryset(request)  # respeita filtro do usuário!
            return exportar_categoria_xlsx(self, request, queryset)
        return super().changelist_view(request, extra_context)

    #list_filter = ('tipo', 'clinica')
    search_fields = ('codigo', 'descricao')
    ordering = ('clinica__nome', 'tipo', 'codigo')

    readonly_fields = []

    fieldsets = (
        ('Identificação', {
            'fields': ('clinica', 'codigo')
        }),
        ('Informações', {
            'fields': ('descricao', 'tipo')
        }),
    )

    # ---------------------------------------------
    # 1. Filtrar por usuário logado (multi-tenant)
    # ---------------------------------------------
    def get_queryset(self, request):
        qs = super().get_queryset(request)

        if request.user.is_superuser:
            return qs

        return qs.filter(clinica__user=request.user)

    # ---------------------------------------------
    # 2. Dropdown de clínica só mostra as do consultor
    # ---------------------------------------------
    def formfield_for_foreignkey(self, db_field, request, **kwargs):

        if db_field.name == "clinica":
            if request.user.is_superuser:
                kwargs["queryset"] = Clinica.objects.all()
            else:
                kwargs["queryset"] = Clinica.objects.filter(user=request.user)

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # ---------------------------------------------
    # 3. Impede trocar a clínica após criar o registro
    # ---------------------------------------------
    def get_readonly_fields(self, request, obj=None):
        if obj:
            return ('clinica',)
        return self.readonly_fields

admin.site.register(Categoria, CategoriaAdmin)
#=============================================








# ===========================================================
#   MIXIN UNIVERSAL — Filtra tudo pela Clínica do consultor
# ===========================================================

class ClinicaFilterMixin:
    """Filtro universal para isolar dados por consultor (user)."""

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(clinica__user=request.user)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        user = request.user

        # FK direto para Clínica
        if db_field.name == "clinica" and not user.is_superuser:
            kwargs["queryset"] = Clinica.objects.filter(user=user)

        # TipoSala → só da clínica do consultor
        if db_field.name == "tipo" and not user.is_superuser:
            kwargs["queryset"] = TipoSala.objects.filter(clinica__user=user)

        # Equipamentos → só da clínica do consultor
        if db_field.name == "equipamentos" and not user.is_superuser:
            kwargs["queryset"] = Equipamento.objects.filter(clinica__user=user)

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def formfield_for_manytomany(self, db_field, request, **kwargs):
        user = request.user

        if db_field.name == "equipamentos" and not user.is_superuser:
            kwargs["queryset"] = Equipamento.objects.filter(clinica__user=user)

        return super().formfield_for_manytomany(db_field, request, **kwargs)



# ===========================================================
#   INLINES
# ===========================================================

class EquipamentoInline(admin.TabularInline):
    model = Sala.equipamentos.through  # MANY-TO-MANY through table
    extra = 1

    def get_formset(self, request, obj=None, **kwargs):
        """
        Captura a Sala que está sendo editada.
        """
        self.sala_obj = obj
        return super().get_formset(request, obj, **kwargs)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """
        Filtra apenas os equipamentos da clínica da Sala atual.
        """
        if db_field.name == "equipamento":

            if hasattr(self, "sala_obj") and self.sala_obj:
                clinica = self.sala_obj.clinica
                kwargs["queryset"] = Equipamento.objects.filter(clinica=clinica)
            else:
                # Criando nova Sala → não existe clínica ainda
                kwargs["queryset"] = Equipamento.objects.none()

        return super().formfield_for_foreignkey(db_field, request, **kwargs)



class CustoFixoInline(admin.TabularInline):
    model = CustoFixoSala
    extra = 1


class CustoVariavelInline(admin.TabularInline):
    model = CustoVariavelSala
    extra = 1


# ===========================================================
#   ADMIN: TipoSala
# ===========================================================

@admin.register(TipoSala)
class TipoSalaAdmin(ClinicaFilterMixin, admin.ModelAdmin):

    list_display = ("tipo", "nome", "clinica", "status")
    #list_filter = ("status", "clinica")
    search_fields = ("tipo", "nome")
    ordering = ("tipo",)

    fieldsets = (
        ("Identificação", {
            "fields": ("clinica", "tipo", "nome", "status")
        }),
        ("Informações", {
            "fields": ("observacao",),
        }),
    )


# Expotar para Excel: Model Parametro
#====================================

def exportar_equipamento_xlsx(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipamentos"

    # Cabeçalho
    ws.append(["Codigo", "Descrição", "Clínica"])

    # Linhas

    for obj in queryset:
        ws.append([obj.codigo, obj.descricao, obj.clinica.nome])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="equipamentos.xlsx"'

    wb.save(response)

    return response

exportar_equipamento_xlsx.short_description = "Export Excel"
#================================================================

#@admin.register(Equipamento)
class EquipamentoAdmin(ClinicaFilterMixin, admin.ModelAdmin):

    list_display = ("codigo", "descricao", "fabricante", "status", "clinica", "custo_dia", "custo_mes")
    #list_filter = ("status", "clinica")
    search_fields = ("codigo", "descricao", "fabricante", "marca")
    ordering = ("codigo",)

    actions = [exportar_equipamento_xlsx]  # adiciona action normal
    change_list_template = "admin/equipamentos_xls.html"

# ---------------------------------------------
# 4. Exporta TODOS os registros da lista (sem precisar selecionar)
# ---------------------------------------------
    def get_actions(self, request):
        actions = super().get_actions(request)
        # Remove a action padrão do admin (que exige seleção)
        if 'exportar_equipamento_xlsx' in actions:
            del actions['exportar_equipamento_xlsx']
        return actions

    # Força a action a sempre usar todos os itens da queryset (a mágica!)
    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'exportar_equipamento_xlsx':
            # Pega TODOS os objetos que o usuário pode ver (com filtro de clínica!)
            queryset = self.get_queryset(request)
            return exportar_equipamento_xlsx(self, request, queryset)
        return super().changelist_view(request, extra_context)

    fieldsets = (
        ("Identificação", {
            "fields": ("clinica", "codigo", "descricao", "status")
        }),
        ("Informações", {
            "fields": ("fabricante", "marca", "dta_aquisicao"),
        }),
        ("Custos", {
            "fields": ("vlr_aquisicao", "custo_dia", "custo_mes"),
        }),
    )
admin.site.register(Equipamento, EquipamentoAdmin)
#=================================================


# ===========================================================
#   ADMIN: Sala (com 2 inlines + M2M filtrado)
# ===========================================================

@admin.register(Sala)
class SalaAdmin(ClinicaFilterMixin, admin.ModelAdmin):

    autocomplete_fields = ('equipamentos',)

    inlines = [CustoFixoInline, CustoVariavelInline]

    list_display = ("numero", "nome", "tipo", "clinica", "status")
    #list_filter = ("status", "clinica", "tipo")
    search_fields = ("numero", "nome")
    ordering = ("numero",)

    fieldsets = (
        ("Identificação", {
            "fields": ("clinica", "numero", "nome", "status")
        }),
        ("Configurações", {
            "fields": ("tipo", "cor"),
        }),
        ("Equipamentos", {
            "fields": ("equipamentos",),
        }),
    )
    class SalaAdmin(admin.ModelAdmin):
        inlines = [EquipamentoInline]



# ===========================================================
#   ADMIN: Custos Fixos & Variáveis (caso precise editar fora da Sala)
# ===========================================================

@admin.register(CustoFixoSala)
class CustoFixoSalaAdmin(admin.ModelAdmin):

    list_display = ("nome_item", "valor_mensal", "sala", "get_clinica", "get_mes", "get_ano")
    #list_filter = ("mes_referencia", "ano_referencia", "sala__clinica")
    search_fields = ("nome_item", "sala__numero")
    ordering = ("ano_referencia", "mes_referencia", "nome_item")

    # FILTRA por clínica do consultor
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(sala__clinica__user=request.user)

    # mostrar a clínica na list_display
    def get_clinica(self, obj):
        return obj.sala.clinica.nome,
    get_clinica.short_description = "Clínica"

    def get_mes(self, obj):
            return obj.mes_referencia
    get_mes.short_description = "Mes Ref."

    def get_ano(self, obj):
            return obj.ano_referencia
    get_ano.short_description = "Ano Ref."
    
    # dropdown de Sala só mostra as Salas do consultor
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "sala" and not request.user.is_superuser:
            kwargs["queryset"] = Sala.objects.filter(clinica__user=request.user)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(CustoVariavelSala)
class CustoVariavelSalaAdmin(admin.ModelAdmin):

    list_display = ("nome_item", "valor_mensal", "sala", "get_clinica", "get_mes", "get_ano")
    #list_filter = ("mes_referencia", "ano_referencia", "sala__clinica")
    search_fields = ("nome_item", "sala__numero")
    ordering = ("ano_referencia", "mes_referencia", "nome_item")

    # FILTRA por clínica do consultor
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(sala__clinica__user=request.user)

    # mostrar a clínica na list_display
    def get_clinica(self, obj):
        return obj.sala.clinica.nome,
    get_clinica.short_description = "Clínica"

    def get_mes(self, obj):
            return obj.mes_referencia
    get_mes.short_description = "Mes Ref."

    def get_ano(self, obj):
            return obj.ano_referencia
    get_ano.short_description = "Ano Ref."
    
    # dropdown de Sala só mostra as Salas do consultor
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "sala" and not request.user.is_superuser:
            kwargs["queryset"] = Sala.objects.filter(clinica__user=request.user)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)






#============ Novos Admin

# ====================== ITEMCUSTO INLINE (lindo e funcional) ======================
class ItemCustoInline(admin.TabularInline):
    model = ItemCusto
    extra = 1
    autocomplete_fields = ['categoria_custo', 'referencia_parametro']
    fields = ('categoria_custo', 'nome', 'quantidade', 'unidade_medida', 'custo_unitario', 'referencia_parametro')
    readonly_fields = ('custo_total_calculado',)

    def custo_total_calculado(self, obj):
        if obj.pk:
            total = obj.calcular_custo_total() if hasattr(obj, 'calcular_custo_total') else obj.quantidade * obj.custo_unitario
            return format_html(f'<b style="color:#28a745;">R$ {total:,.2f}</b>')
        return "-"
    custo_total_calculado.short_description = "Custo Total"


# ====================== PROCEDIMENTO ======================
# clinicas/admin.py
from django.contrib import admin
from django.utils.html import format_html
from .models import * 


# ====================== FUNÇÃO MÁGICA QUE RESOLVE TUDO ======================
def get_clinica_queryset(request):
    """Retorna queryset de clínicas filtradas por usuário (com edição funcionando)"""
    if request.user.is_superuser:
        return Clinica.objects.all()
    
    qs = Clinica.objects.filter(user=request.user)
    
    # Se estiver editando um objeto, garante que a clínica atual apareça no dropdown
    if hasattr(request, 'resolver_match'):
        obj_id = request.resolver_match.kwargs.get('object_id')
        if obj_id:
            try:
                # Tenta pegar do objeto sendo editado (qualquer model que tenha clinica)
                from django.apps import apps
                for model in [Procedimento, Convenio, TabelaPreco, HistoricoPreco, ItemCusto]:
                    try:
                        obj = model.objects.get(pk=obj_id)
                        if hasattr(obj, 'clinica') and obj.clinica not in qs:
                            qs = qs | Clinica.objects.filter(pk=obj.clinica.pk)
                            break
                    except:
                        continue
            except:
                pass
    return qs.distinct()

"""
# ====================== FUNÇÃO MÁGICA QUE RESOLVE TUDO ======================
def get_categoria_queryset(request):
    #Retorna queryset de categorias filtradas por usuário (com edição funcionando)
    if request.user.is_superuser:
        return Categoria.objects.all()
    
    qs = Categoria.objects.filter(user=request.user)
    
    # Se estiver editando um objeto, garante que a Categoria atual apareça no dropdown
    if hasattr(request, 'resolver_match'):
        obj_id = request.resolver_match.kwargs.get('object_id')
        if obj_id:
            try:
                # Tenta pegar do objeto sendo editado (qualquer model que tenha Categoria)
                from django.apps import apps
                for model in [Procedimento, Convenio, TabelaPreco, HistoricoPreco, ItemCusto]:
                    try:
                        obj = model.objects.get(pk=obj_id)
                        if hasattr(obj, 'categoria') and obj.categoria not in qs:
                            qs = qs | Categoria.objects.filter(pk=obj.categoria.pk)
                            break
                    except:
                        continue
            except:
                pass
    return qs.distinct()
"""

# ====================== PROCEDIMENTO ======================
@admin.register(Procedimento)
class ProcedimentoAdmin(admin.ModelAdmin):
    list_display = ('codigo', 'nome', 'clinica_nome', 'categoria', 'status')
    list_filter = ('clinica__nome', 'categoria', 'status')
    search_fields = ('codigo', 'nome', 'clinica__nome')
    ordering = ('clinica__nome', 'nome')

    def clinica_nome(self, obj):
        return obj.clinica.nome
    clinica_nome.short_description = "Clínica"
    clinica_nome.admin_order_field = 'clinica__nome'

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "clinica":
            kwargs["queryset"] = get_clinica_queryset(request)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)



# ====================== CONVÊNIO ======================
@admin.register(Convenio)
class ConvenioAdmin(admin.ModelAdmin):
    list_display = ('nome', 'clinica_nome', 'fator_reajuste', 'data_atualizacao')
    #list_filter = ('clinica__nome',)
    search_fields = ('nome', 'clinica__nome')

    def clinica_nome(self, obj):
        return obj.clinica.nome
    clinica_nome.short_description = "Clínica"

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "clinica":
            kwargs["queryset"] = get_clinica_queryset(request)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


# ====================== TABELA DE PREÇO ======================
@admin.register(TabelaPreco)
class TabelaPrecoAdmin(admin.ModelAdmin):
    list_display = ('procedimento', 'convenio', 'clinica_nome', 'preco_venda', 'margem_efetiva', 'status_margem')
    list_filter = ('clinica__nome', 'convenio__nome', 'status_margem')
    search_fields = ('procedimento__nome', 'convenio__nome', 'clinica__nome')

    def clinica_nome(self, obj):
        return obj.clinica.nome
    clinica_nome.short_description = "Clínica"

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "clinica":
            kwargs["queryset"] = get_clinica_queryset(request)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


# ====================== ITEM CUSTO ======================
@admin.register(ItemCusto)
class ItemCustoAdmin(admin.ModelAdmin):
    list_display = ('procedimento', 'nome', 'clinica_nome', 'quantidade', 'custo_total_item')
    list_filter = ('clinica__nome', 'categoria_custo')
    search_fields = ('nome', 'procedimento__nome')

    def clinica_nome(self, obj):
        return obj.clinica.nome
    clinica_nome.short_description = "Clínica"

    def custo_total_item(self, obj):
        total = obj.quantidade * obj.custo_unitario
        return f"R$ {total:,.2f}"
    custo_total_item.short_description = "Custo Total"

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "clinica":
            kwargs["queryset"] = get_clinica_queryset(request)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


# ====================== HISTÓRICO PREÇO ======================
@admin.register(HistoricoPreco)
class HistoricoPrecoAdmin(admin.ModelAdmin):
    list_display = ('tabela_preco', 'clinica_nome', 'preco_venda_anterior', 'data_alteracao')
    list_filter = ('clinica__nome', 'data_alteracao')
    readonly_fields = ('data_alteracao',)

    def clinica_nome(self, obj):
        return obj.clinica.nome
    clinica_nome.short_description = "Clínica"

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "clinica":
            kwargs["queryset"] = get_clinica_queryset(request)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)